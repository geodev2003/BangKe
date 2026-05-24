from fastapi import FastAPI, UploadFile, File, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from pydantic import BaseModel
from collections import defaultdict
import tempfile, os

from database import get_db, init_db
from models import ServiceGroup, Service, Patient, OutpatientBill, BillItem
from excel_parser import parse_excel
from word_generator import generate_word, generate_pdf

def calc_bhyt(so_luong, don_gia_bv, don_gia_bh, ty_le_tt_dv, ty_le_bhyt):
    """
    Tính các cột BHYT cho 1 dòng dịch vụ.
    ty_le_bhyt: mức hưởng BHYT (0-100), None nếu không có BHYT.
    Returns: (thanh_tien_bv, thanh_tien_bh, quy_bhyt, nb_cung_tt, nb_tu_tra)
    """
    sl  = so_luong or 1
    dgbv = don_gia_bv or 0
    dgbh = don_gia_bh or 0
    tl_dv = ty_le_tt_dv or 100

    tt_bv = sl * dgbv  # thành tiền BV

    if ty_le_bhyt and ty_le_bhyt > 0 and dgbh > 0:
        # Thành tiền BH = SL × đơn giá BH × tỷ lệ TT DV%
        tt_bh_base = sl * dgbh * tl_dv / 100
        # Quỹ BHYT = thành tiền BH × mức hưởng%
        quy = round(tt_bh_base * ty_le_bhyt / 100)
        tt_bh = round(tt_bh_base)
        nb_cung = round(tt_bh_base - quy)   # NB cùng chi trả
        nb_tu   = round(tt_bv - quy)         # NB tự trả = TT BV - quỹ
        return tt_bv, tt_bh, quy, nb_cung, nb_tu
    else:
        # Không có BHYT
        return tt_bv, None, None, None, tt_bv

app = FastAPI(title="Bảng Kê Ngoại Trú API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def on_startup():
    init_db()

GROUP_PREFIX_MAP = {
    'Xét nghiệm':             '1. ',
    'Chẩn đoán hình ảnh':     '2. ',
    'Thăm dò chức năng':      '3. ',
    'Phẫu thuật - Thủ thuật': '4. ',
    'Khám bệnh':              '5. ',
    'Thuốc':                  '6. ',
}

# ── Schemas ───────────────────────────────────────────────────────────────────
class ServiceGroupCreate(BaseModel):
    name: str
    display_order: int = 0

class ServiceCreate(BaseModel):
    name: str
    unit: str = "Lần"
    price: Optional[float] = None
    bhyt_price: Optional[float] = None
    group_id: int
    display_order: int = 0

class ServiceUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    bhyt_price: Optional[float] = None
    group_id: Optional[int] = None
    display_order: Optional[int] = None

class BillItemAdd(BaseModel):
    service_id: int
    so_luong: float = 1

class BillItemUpdate(BaseModel):
    so_luong: Optional[float] = None
    don_gia_bv: Optional[float] = None
    don_gia_bh: Optional[float] = None
    ty_le_bhyt: Optional[float] = None   # mức hưởng BHYT %

class BillBHYTUpdate(BaseModel):
    """Cập nhật mức hưởng BHYT cho toàn bộ bảng kê."""
    ty_le_bhyt: Optional[float] = None        # None = xóa BHYT, 0-100 = %
    don_gia_bh_map: Optional[dict] = None     # {item_id: don_gia_bh}
    excluded_item_ids: Optional[List[int]] = None  # item IDs không áp dụng BHYT

# ── Service Groups ─────────────────────────────────────────────────────────────
@app.get("/api/service-groups")
def list_groups(db: Session = Depends(get_db)):
    groups = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    return [{
        "id": g.id, "name": g.name, "display_order": g.display_order,
        "services": [{
            "id": s.id, "name": s.name, "unit": s.unit,
            "price": s.price, "bhyt_price": s.bhyt_price,
            "group_id": s.group_id, "display_order": s.display_order,
        } for s in sorted(g.services, key=lambda x: x.display_order)]
    } for g in groups]

@app.post("/api/service-groups")
def create_group(body: ServiceGroupCreate, db: Session = Depends(get_db)):
    grp = ServiceGroup(name=body.name, display_order=body.display_order)
    db.add(grp); db.commit(); db.refresh(grp)
    return {"id": grp.id, "name": grp.name, "display_order": grp.display_order, "services": []}

@app.delete("/api/service-groups/{group_id}")
def delete_group(group_id: int, db: Session = Depends(get_db)):
    grp = db.query(ServiceGroup).filter(ServiceGroup.id == group_id).first()
    if not grp: raise HTTPException(404, "Không tìm thấy nhóm")
    db.delete(grp); db.commit()
    return {"ok": True}

@app.post("/api/services")
def create_service(body: ServiceCreate, db: Session = Depends(get_db)):
    svc = Service(**body.dict())
    db.add(svc); db.commit(); db.refresh(svc)
    return {"id": svc.id, "name": svc.name, "unit": svc.unit, "price": svc.price, "bhyt_price": svc.bhyt_price, "group_id": svc.group_id}

@app.put("/api/services/{service_id}")
def update_service(service_id: int, body: ServiceUpdate, db: Session = Depends(get_db)):
    svc = db.query(Service).filter(Service.id == service_id).first()
    if not svc: raise HTTPException(404, "Không tìm thấy dịch vụ")
    for k, v in body.dict(exclude_none=True).items():
        setattr(svc, k, v)
    db.commit(); db.refresh(svc)
    return {"id": svc.id, "name": svc.name, "unit": svc.unit,
            "price": svc.price, "bhyt_price": svc.bhyt_price}

@app.delete("/api/services/{service_id}")
def delete_service(service_id: int, db: Session = Depends(get_db)):
    svc = db.query(Service).filter(Service.id == service_id).first()
    if not svc: raise HTTPException(404, "Không tìm thấy dịch vụ")
    db.delete(svc); db.commit()
    return {"ok": True}

# ── Build bill data ────────────────────────────────────────────────────────────
def _build_groups(bill_items, groups_db):
    items_by_group = defaultdict(list)
    for item in bill_items:
        if item.service:
            tt_bv = item.thanh_tien_bv or 0
            items_by_group[item.service.group_id].append({
                "id":            item.id,
                "name":          item.service.name,
                "unit":          item.service.unit,
                "so_luong":      item.so_luong,
                "don_gia_bv":    item.don_gia_bv or 0,
                "don_gia_bh":    item.don_gia_bh or 0,
                "ty_le_dv":      item.ty_le_tt_dv or 100,
                "thanh_tien_bv": tt_bv,
                "ty_le_bh":      item.ty_le_bhyt or "",
                "thanh_tien_bh": item.thanh_tien_bh,
                "quy_bhyt":      item.quy_bhyt,
                "nb_cung_tt":    item.nb_cung_tt,
                "khac":          None,
                "nb_tu_tra":     item.nb_tu_tra if item.nb_tu_tra is not None else tt_bv,
            })
    result = []
    for g in groups_db:
        if g.id in items_by_group:
            items = items_by_group[g.id]
            grp_bv  = sum(i["thanh_tien_bv"] or 0 for i in items)
            grp_bh  = sum(i["thanh_tien_bh"] or 0 for i in items) or None
            grp_quy = sum(i["quy_bhyt"] or 0 for i in items) or None
            grp_nb  = sum(i["nb_tu_tra"] or 0 for i in items)
            result.append({
                "group_name":     g.name,
                "group_prefix":   GROUP_PREFIX_MAP.get(g.name, ''),
                "items":          items,
                "tong_tt_bv":     grp_bv,
                "tong_bh":        grp_bh,
                "tong_quy":       grp_quy,
                "tong_nb_tu_tra": grp_nb,
            })
    return result


def _bill_to_dict(bill, groups_db):
    import json
    p = bill.patient
    groups_result = _build_groups(bill.items, groups_db)
    total = sum(i.thanh_tien_bv or 0 for i in bill.items)
    # Read extra fields stored as JSON in ten_goi_kham
    extra = {}
    ten_goi = p.ten_goi_kham or ""
    if ten_goi.startswith('{'):
        try: extra = json.loads(ten_goi)
        except: pass
    else:
        extra["ten_goi_kham_display"] = ten_goi
    return {
        "id": bill.id,
        "ma_bn": p.ma_bn,
        "so_kham_benh": "",
        "khoa": extra.get("khoa", "Khoa Khám Bệnh"),
        "ho_ten": p.ho_ten,
        "gioi_tinh": p.gioi_tinh,
        "ngay_sinh": p.ngay_sinh,
        "dia_chi": p.dia_chi,
        "ma_khu_vuc": "",
        "ma_the_bhyt": extra.get("ma_the_bhyt", ""),
        "bhyt_tu": extra.get("bhyt_tu", ""),
        "bhyt_den": extra.get("bhyt_den", ""),
        "csdk_bhyt": extra.get("csdk_bhyt", ""),
        "ma_csdk": "",
        "ngay_kham": bill.ngay_kham or "",
        "chan_doan": extra.get("chan_doan", ""),
        "ma_benh": extra.get("ma_benh", ""),
        "benh_kem_theo": "", "ma_benh_kem": "",
        "muc_huong_bhyt": extra.get("muc_huong_bhyt", ""),
        "ngay_tinh_tu": "", "ngay_tinh_den": "",
        "ten_goi_kham": extra.get("ten_goi_kham_display", ""),
        "created_at": bill.created_at.isoformat(),
        "groups": groups_result,
        "tong_cong_bv": total, "tong_bh": None, "tong_quy": None,
        "tong_nb_tu_tra": total, "tong_chi_phi": total,
        "quy_bhyt_tt": 0, "nb_tu_tra": total, "khac_tt": 0, "total": total,
    }

def build_bill_data(patient_row: dict, db: Session) -> dict:
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    groups_result = []
    tong_cong = 0.0
    for grp in groups_db:
        items = []
        for svc in sorted(grp.services, key=lambda x: x.display_order):
            qty = next((q for n, q in patient_row["services"].items()
                        if n.strip().lower() == svc.name.strip().lower()), None)
            if qty and qty > 0:
                price = svc.price or 0
                tt = price * qty
                tong_cong += tt
                items.append({
                    "name": svc.name.strip(), "unit": svc.unit,
                    "so_luong": qty, "don_gia_bv": price,
                    "don_gia_bh": 0, "ty_le_dv": 100, "thanh_tien_bv": tt,
                    "ty_le_bh": "", "thanh_tien_bh": None,
                    "quy_bhyt": None, "nb_cung_tt": None, "khac": None,
                    "nb_tu_tra": tt,
                })
        if items:
            grp_total = sum(i["thanh_tien_bv"] for i in items)
            groups_result.append({
                "group_name": grp.name,
                "group_prefix": GROUP_PREFIX_MAP.get(grp.name, ''),
                "items": items,
                "tong_tt_bv": grp_total,
                "tong_bh": None, "tong_quy": None, "tong_nb_tu_tra": grp_total,
            })
    return {
        "ma_bn": patient_row.get("ma_bn",""), "so_kham_benh": "", "khoa": "Khoa Khám Bệnh",
        "ho_ten": patient_row.get("ho_ten",""), "ngay_sinh": patient_row.get("ngay_sinh",""),
        "gioi_tinh": patient_row.get("gioi_tinh",""), "dia_chi": patient_row.get("dia_chi",""),
        "ma_khu_vuc":"","ma_the_bhyt":"","bhyt_tu":"","bhyt_den":"",
        "csdk_bhyt":"","ma_csdk":"","ngay_kham": patient_row.get("ngay_kham",""),
        "chan_doan":"","ma_benh":"","benh_kem_theo":"","ma_benh_kem":"",
        "muc_huong_bhyt":"","ngay_tinh_tu":"","ngay_tinh_den":"",
        "ten_goi_kham": patient_row.get("ten_goi_kham",""),
        "groups": groups_result,
        "tong_cong_bv": tong_cong, "tong_bh": None, "tong_quy": None,
        "tong_nb_tu_tra": tong_cong, "tong_chi_phi": tong_cong,
        "quy_bhyt_tt": 0, "nb_tu_tra": tong_cong, "khac_tt": 0, "total": tong_cong,
    }

def save_bill_to_db(patient_row, bill_data, db):
    patient = db.query(Patient).filter(Patient.ma_bn == bill_data["ma_bn"]).first()
    if not patient:
        patient = Patient(
            ma_bn=bill_data["ma_bn"], ho_ten=bill_data["ho_ten"],
            gioi_tinh=bill_data["gioi_tinh"], ngay_sinh=bill_data["ngay_sinh"],
            dia_chi=bill_data["dia_chi"], ten_goi_kham=bill_data["ten_goi_kham"],
        )
        db.add(patient); db.flush()
    else:
        patient.ho_ten = bill_data["ho_ten"]
        patient.dia_chi = bill_data["dia_chi"]

    bill = OutpatientBill(patient_id=patient.id, ngay_kham=bill_data["ngay_kham"])
    db.add(bill); db.flush()

    svc_map = {s.name.strip().lower(): s for s in db.query(Service).all()}
    for grp in bill_data["groups"]:
        for item in grp["items"]:
            svc_obj = svc_map.get(item["name"].strip().lower())
            svc_id  = svc_obj.id if svc_obj else None
            # Ưu tiên: 1) giá BH từ item, 2) bhyt_price từ danh mục, 3) 0
            dgbh = item.get("don_gia_bh") or (svc_obj.bhyt_price if svc_obj else 0) or 0
            tt_bv, _, _, _, nb_tu = calc_bhyt(
                item["so_luong"], item["don_gia_bv"], dgbh,
                item.get("ty_le_dv",100), None
            )
            db.add(BillItem(
                bill_id=bill.id,
                service_id=svc_id,
                so_luong=item["so_luong"],
                don_gia_bv=item["don_gia_bv"],
                don_gia_bh=dgbh,
                ty_le_tt_dv=item.get("ty_le_dv",100),
                thanh_tien_bv=tt_bv,
                nb_tu_tra=nb_tu,
            ))
    db.commit()
    return bill

# ── Upload Excel ───────────────────────────────────────────────────────────────
@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        patients = parse_excel(tmp_path)
        results = []
        for p in patients:
            bill_data = build_bill_data(p, db)
            bill = save_bill_to_db(p, bill_data, db)
            results.append({
                "ma_bn": bill_data["ma_bn"], "ho_ten": bill_data["ho_ten"],
                "bill_id": bill.id, "total": bill_data["total"],
                "services_count": sum(len(g["items"]) for g in bill_data["groups"]),
            })
        return {"success": True, "patients": results, "count": len(results)}
    finally:
        os.unlink(tmp_path)

# ── Bills ──────────────────────────────────────────────────────────────────────
@app.get("/api/bills")
def list_bills(db: Session = Depends(get_db)):
    bills = db.query(OutpatientBill).join(Patient)\
               .order_by(OutpatientBill.created_at.desc()).all()
    return [{
        "id": b.id, "ma_bn": b.patient.ma_bn, "ho_ten": b.patient.ho_ten,
        "gioi_tinh": b.patient.gioi_tinh, "ngay_sinh": b.patient.ngay_sinh,
        "ngay_kham": b.ngay_kham, "created_at": b.created_at.isoformat(),
        "total": sum(i.thanh_tien_bv or 0 for i in b.items),
        "items_count": len(b.items),
    } for b in bills]

@app.get("/api/bills/{bill_id}")
def get_bill(bill_id: int, db: Session = Depends(get_db)):
    bill = db.query(OutpatientBill).filter(OutpatientBill.id == bill_id).first()
    if not bill: raise HTTPException(404, "Không tìm thấy bảng kê")
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    return _bill_to_dict(bill, groups_db)

@app.delete("/api/bills/{bill_id}")
def delete_bill(bill_id: int, db: Session = Depends(get_db)):
    bill = db.query(OutpatientBill).filter(OutpatientBill.id == bill_id).first()
    if not bill: raise HTTPException(404, "Không tìm thấy bảng kê")
    db.delete(bill); db.commit()
    return {"ok": True}

# ── Thêm / Xóa / Sửa dịch vụ trong bảng kê ───────────────────────────────────
@app.post("/api/bills/{bill_id}/items")
def add_item_to_bill(bill_id: int, body: BillItemAdd, db: Session = Depends(get_db)):
    """Thêm một dịch vụ vào bảng kê đã tạo."""
    bill = db.query(OutpatientBill).filter(OutpatientBill.id == bill_id).first()
    if not bill: raise HTTPException(404, "Không tìm thấy bảng kê")

    svc = db.query(Service).filter(Service.id == body.service_id).first()
    if not svc: raise HTTPException(404, "Không tìm thấy dịch vụ")

# Kiểm tra dịch vụ đã có chưa
    existing = next((i for i in bill.items if i.service_id == body.service_id), None)
    if existing:
        existing.so_luong += body.so_luong
        tt_bv, tt_bh, quy, nb_cung, nb_tu = calc_bhyt(
            existing.so_luong, existing.don_gia_bv, existing.don_gia_bh,
            existing.ty_le_tt_dv, existing.ty_le_bhyt
        )
        existing.thanh_tien_bv = tt_bv
        existing.thanh_tien_bh = tt_bh
        existing.quy_bhyt = quy
        existing.nb_cung_tt = nb_cung
        existing.nb_tu_tra = nb_tu
    else:
        price      = svc.price or 0
        bhyt_price = svc.bhyt_price or 0  # lấy giá BHYT từ danh mục dịch vụ
        tt_bv, tt_bh, quy, nb_cung, nb_tu = calc_bhyt(body.so_luong, price, bhyt_price, 100, None)
        db.add(BillItem(
            bill_id=bill.id, service_id=svc.id,
            so_luong=body.so_luong, don_gia_bv=price,
            don_gia_bh=bhyt_price, ty_le_tt_dv=100,
            thanh_tien_bv=tt_bv, nb_tu_tra=nb_tu,
        ))
    db.commit()
    db.refresh(bill)
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    return _bill_to_dict(bill, groups_db)

@app.put("/api/bills/{bill_id}/items/{item_id}")
def update_bill_item(bill_id: int, item_id: int, body: BillItemUpdate, db: Session = Depends(get_db)):
    """Sửa số lượng, đơn giá hoặc BHYT của một dịch vụ trong bảng kê."""
    item = db.query(BillItem).filter(
        BillItem.id == item_id, BillItem.bill_id == bill_id
    ).first()
    if not item: raise HTTPException(404, "Không tìm thấy dòng dịch vụ")

    if body.so_luong  is not None: item.so_luong  = body.so_luong
    if body.don_gia_bv is not None: item.don_gia_bv = body.don_gia_bv
    if body.don_gia_bh is not None: item.don_gia_bh = body.don_gia_bh
    if body.ty_le_bhyt is not None: item.ty_le_bhyt = body.ty_le_bhyt

    tt_bv, tt_bh, quy, nb_cung, nb_tu = calc_bhyt(
        item.so_luong, item.don_gia_bv, item.don_gia_bh,
        item.ty_le_tt_dv, item.ty_le_bhyt
    )
    item.thanh_tien_bv = tt_bv
    item.thanh_tien_bh = tt_bh
    item.quy_bhyt      = quy
    item.nb_cung_tt    = nb_cung
    item.nb_tu_tra     = nb_tu

    db.commit()
    bill = db.query(OutpatientBill).filter(OutpatientBill.id == bill_id).first()
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    return _bill_to_dict(bill, groups_db)

@app.delete("/api/bills/{bill_id}/items/{item_id}")
def delete_bill_item(bill_id: int, item_id: int, db: Session = Depends(get_db)):
    """Xóa một dịch vụ khỏi bảng kê."""
    item = db.query(BillItem).filter(
        BillItem.id == item_id, BillItem.bill_id == bill_id
    ).first()
    if not item: raise HTTPException(404, "Không tìm thấy dòng dịch vụ")
    db.delete(item); db.commit()
    return {"ok": True}

@app.put("/api/bills/{bill_id}/bhyt")
def update_bill_bhyt(bill_id: int, body: BillBHYTUpdate, db: Session = Depends(get_db)):
    """Áp dụng mức hưởng BHYT cho toàn bộ bảng kê."""
    bill = db.query(OutpatientBill).filter(OutpatientBill.id == bill_id).first()
    if not bill: raise HTTPException(404, "Không tìm thấy bảng kê")

    ty_le = body.ty_le_bhyt  # None = xóa BHYT
    dgbh_map = body.don_gia_bh_map or {}

    excluded = set(body.excluded_item_ids or [])
    dgbh_map = body.don_gia_bh_map or {}

    for item in bill.items:
        if not item.service: continue
        # Nếu item bị loại trừ → xóa BHYT cho item này
        if item.id in excluded:
            item.don_gia_bh  = 0
            item.ty_le_bhyt  = None
            item.thanh_tien_bh = None
            item.quy_bhyt    = None
            item.nb_cung_tt  = None
            item.nb_tu_tra   = item.thanh_tien_bv  # NB tự trả = toàn bộ
            continue

        # Lấy đơn giá BH từ map theo item.id, fallback service_id, fallback đơn giá BV
        # Priority: 1) from form map, 2) existing item don_gia_bh, 3) service bhyt_price, 4) don_gia_bv
        svc_bhyt = item.service.bhyt_price if item.service else 0
        dgbh = dgbh_map.get(str(item.id),
               dgbh_map.get(str(item.service_id),
               item.don_gia_bh or svc_bhyt or item.don_gia_bv or 0))
        item.don_gia_bh = float(dgbh) if dgbh != '' else (item.don_gia_bv or 0)
        item.ty_le_bhyt = ty_le

        tt_bv, tt_bh, quy, nb_cung, nb_tu = calc_bhyt(
            item.so_luong, item.don_gia_bv, item.don_gia_bh,
            item.ty_le_tt_dv or 100, ty_le
        )
        item.thanh_tien_bv = tt_bv
        item.thanh_tien_bh = tt_bh
        item.quy_bhyt      = quy
        item.nb_cung_tt    = nb_cung
        item.nb_tu_tra     = nb_tu

    db.commit()
    db.refresh(bill)
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    return _bill_to_dict(bill, groups_db)


# ── Export ─────────────────────────────────────────────────────────────────────
@app.get("/api/bills/{bill_id}/export/word")
def export_word(bill_id: int, db: Session = Depends(get_db)):
    bill_data = get_bill(bill_id, db)
    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        out = tmp.name
    try:
        generate_word(bill_data, out)
        fname = f"bang_ke_{bill_data['ma_bn']}_{bill_data['ho_ten'].replace(' ','_')}.docx"
        return FileResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename=fname,
        )
    except Exception as e:
        os.unlink(out)
        raise HTTPException(500, str(e))

@app.get("/api/bills/{bill_id}/export/pdf")
def export_pdf(bill_id: int, db: Session = Depends(get_db)):
    bill_data = get_bill(bill_id, db)
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        out = tmp.name
    try:
        generate_pdf(bill_data, out)
        fname = f"bang_ke_{bill_data['ma_bn']}_{bill_data['ho_ten'].replace(' ','_')}.pdf"
        return FileResponse(out, media_type="application/pdf", filename=fname)
    except Exception as e:
        os.unlink(out)
        raise HTTPException(500, str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


# ── Cập nhật thông tin bệnh nhân & bảng kê ────────────────────────────────────
class BillInfoUpdate(BaseModel):
    ho_ten: Optional[str] = None
    ngay_sinh: Optional[str] = None
    gioi_tinh: Optional[str] = None
    dia_chi: Optional[str] = None
    ngay_kham: Optional[str] = None
    chan_doan: Optional[str] = None
    ma_benh: Optional[str] = None
    ma_the_bhyt: Optional[str] = None
    bhyt_tu: Optional[str] = None
    bhyt_den: Optional[str] = None
    muc_huong_bhyt: Optional[str] = None
    csdk_bhyt: Optional[str] = None
    khoa: Optional[str] = None

@app.put("/api/bills/{bill_id}/info")
def update_bill_info(bill_id: int, body: BillInfoUpdate, db: Session = Depends(get_db)):
    """Cập nhật thông tin bệnh nhân và bảng kê."""
    bill = db.query(OutpatientBill).filter(OutpatientBill.id == bill_id).first()
    if not bill: raise HTTPException(404, "Không tìm thấy bảng kê")

    p = bill.patient
    if body.ho_ten    is not None: p.ho_ten    = body.ho_ten
    if body.ngay_sinh is not None: p.ngay_sinh = body.ngay_sinh
    if body.gioi_tinh is not None: p.gioi_tinh = body.gioi_tinh
    if body.dia_chi   is not None: p.dia_chi   = body.dia_chi

    if body.ngay_kham is not None: bill.ngay_kham = body.ngay_kham

    # Lưu extra info vào ten_goi_kham field tạm thời dưới dạng JSON
    import json
    extra = {}
    try:
        if p.ten_goi_kham and p.ten_goi_kham.startswith('{'):
            extra = json.loads(p.ten_goi_kham)
    except: pass

    for field in ['chan_doan','ma_benh','ma_the_bhyt','bhyt_tu','bhyt_den',
                  'muc_huong_bhyt','csdk_bhyt','khoa']:
        val = getattr(body, field)
        if val is not None:
            extra[field] = val

    if extra:
        p.ten_goi_kham = json.dumps(extra, ensure_ascii=False)

    db.commit()
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    return _bill_to_dict(bill, groups_db)


# ── Preview endpoint ──────────────────────────────────────────────────────────
import base64

@app.get("/api/bills/{bill_id}/preview")
def preview_bill(bill_id: int, db: Session = Depends(get_db)):
    """Trả về PDF dạng base64 để hiển thị preview trên trình duyệt."""
    bill_data = get_bill(bill_id, db)
    with tempfile.TemporaryDirectory() as tmp:
        pdf_path = os.path.join(tmp, 'preview.pdf')
        try:
            generate_pdf(bill_data, pdf_path)
            with open(pdf_path, 'rb') as f:
                pdf_b64 = base64.b64encode(f.read()).decode()
            return {
                "pdf_base64": pdf_b64,
                "filename": f"bang_ke_{bill_data['ma_bn']}.pdf"
            }
        except Exception as e:
            raise HTTPException(500, f"Lỗi tạo preview: {str(e)}")


# ── Thống kê ──────────────────────────────────────────────────────────────────
@app.get("/api/stats")
def get_stats(db: Session = Depends(get_db)):
    """Thống kê tổng quan."""
    from sqlalchemy import func
    from models import OutpatientBill, BillItem, Patient

    total_bills  = db.query(OutpatientBill).count()
    total_patients = db.query(Patient).count()
    total_revenue = db.query(func.sum(BillItem.thanh_tien_bv)).scalar() or 0

    # Top 5 dịch vụ được dùng nhiều nhất
    top_svcs = db.query(
        Service.name,
        func.count(BillItem.id).label('count'),
        func.sum(BillItem.thanh_tien_bv).label('revenue'),
    ).join(BillItem, BillItem.service_id == Service.id)\
     .group_by(Service.id)\
     .order_by(func.count(BillItem.id).desc())\
     .limit(5).all()

    # Bills theo ngày (7 ngày gần nhất)
    recent = db.query(
        OutpatientBill.created_at,
        func.sum(BillItem.thanh_tien_bv).label('revenue'),
    ).join(BillItem)\
     .group_by(OutpatientBill.id)\
     .order_by(OutpatientBill.created_at.desc())\
     .limit(20).all()

    return {
        "total_bills": total_bills,
        "total_patients": total_patients,
        "total_revenue": float(total_revenue),
        "avg_per_bill": float(total_revenue / total_bills) if total_bills else 0,
        "top_services": [
            {"name": s.name, "count": s.count, "revenue": float(s.revenue or 0)}
            for s in top_svcs
        ],
        "recent_bills": [
            {"date": b.created_at.strftime("%d/%m"), "revenue": float(b.revenue or 0)}
            for b in recent
        ],
    }

# ── Tìm kiếm bệnh nhân ──────────────────────────────────────────────────────
@app.get("/api/patients/search")
def search_patients(q: str = "", db: Session = Depends(get_db)):
    """Tìm kiếm bệnh nhân theo tên hoặc mã."""
    patients = db.query(Patient).filter(
        (Patient.ho_ten.ilike(f"%{q}%")) |
        (Patient.ma_bn.ilike(f"%{q}%"))
    ).limit(20).all()
    return [{
        "id": p.id, "ma_bn": p.ma_bn, "ho_ten": p.ho_ten,
        "gioi_tinh": p.gioi_tinh, "ngay_sinh": p.ngay_sinh,
        "dia_chi": p.dia_chi,
        "so_lan_kham": len(p.bills),
    } for p in patients]


# ── Xuất nhiều bảng kê ────────────────────────────────────────────────────────
import zipfile
from typing import List as TList

class BulkExportRequest(BaseModel):
    bill_ids: TList[int]
    format: str = "pdf"  # "pdf" or "word"

@app.post("/api/bills/export-bulk")
def export_bulk(body: BulkExportRequest, db: Session = Depends(get_db)):
    """Xuất nhiều bảng kê thành 1 file ZIP."""
    if not body.bill_ids:
        raise HTTPException(400, "Vui lòng chọn ít nhất 1 bảng kê")
    if len(body.bill_ids) > 50:
        raise HTTPException(400, "Tối đa 50 bảng kê mỗi lần xuất")

    with tempfile.TemporaryDirectory() as tmp:
        zip_path = os.path.join(tmp, "bang_ke_export.zip")
        errors = []

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for bill_id in body.bill_ids:
                try:
                    bill_data = get_bill(bill_id, db)
                    ma_bn   = bill_data['ma_bn']
                    ho_ten  = bill_data['ho_ten'].replace(' ', '_')
                    ext     = 'docx' if body.format == 'word' else 'pdf'
                    fname   = f"bang_ke_{ma_bn}_{ho_ten}.{ext}"

                    out = os.path.join(tmp, fname)
                    if body.format == 'word':
                        generate_word(bill_data, out)
                    else:
                        generate_pdf(bill_data, out)

                    zf.write(out, fname)
                except Exception as e:
                    errors.append(f"Bảng kê #{bill_id}: {str(e)}")

        if errors and len(errors) == len(body.bill_ids):
            raise HTTPException(500, "Không thể xuất bất kỳ bảng kê nào: " + "; ".join(errors))

        # Return zip as base64
        with open(zip_path, 'rb') as f:
            zip_b64 = base64.b64encode(f.read()).decode()

        return {
            "zip_base64": zip_b64,
            "filename": f"bang_ke_export_{len(body.bill_ids)}.zip",
            "success_count": len(body.bill_ids) - len(errors),
            "errors": errors,
        }


# ── Import dịch vụ từ Excel ───────────────────────────────────────────────────
@app.post("/api/services/import-excel")
async def import_services_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import dịch vụ từ file Excel.
    Cột bắt buộc: Tên dịch vụ, Nhóm
    Cột tùy chọn: Đơn vị, Đơn giá BV, Đơn giá BHYT, Thứ tự
    """
    import openpyxl
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Tìm header row
        header_idx = None
        header_map = {}
        ALIASES = {
            "ten_dich_vu": ["tên dịch vụ", "ten dich vu", "dịch vụ", "dich vu", "tên", "ten", "name", "service"],
            "nhom":        ["nhóm", "nhom", "nhóm dịch vụ", "group", "loại", "loai"],
            "don_vi":      ["đơn vị", "don vi", "dvt", "đvt", "unit"],
            "don_gia_bv":  ["đơn giá bv", "don gia bv", "giá bv", "gia bv", "đơn giá", "don gia", "giá", "gia", "price"],
            "don_gia_bh":  ["đơn giá bhyt", "don gia bhyt", "giá bhyt", "gia bhyt", "bhyt price", "bhyt_price"],
            "thu_tu":      ["thứ tự", "thu tu", "stt", "order", "số thứ tự"],
        }

        for i, row in enumerate(rows[:5]):
            if not row: continue
            row_lower = [str(c).strip().lower() if c else "" for c in row]
            matched = 0
            tmp_map = {}
            for field, aliases in ALIASES.items():
                for j, cell in enumerate(row_lower):
                    if cell in aliases:
                        tmp_map[field] = j
                        matched += 1
                        break
            if matched >= 2 and "ten_dich_vu" in tmp_map:
                header_idx = i
                header_map = tmp_map
                break

        if header_idx is None or "ten_dich_vu" not in header_map:
            raise HTTPException(400,
                "Không tìm thấy cột 'Tên dịch vụ'. "
                "File cần có các cột: Tên dịch vụ, Nhóm (bắt buộc), "
                "Đơn vị, Đơn giá BV, Đơn giá BHYT (tùy chọn)."
            )

        # Load nhóm hiện có
        groups_db = {g.name.strip().lower(): g for g in db.query(ServiceGroup).all()}
        svcs_db   = {s.name.strip().lower(): s for s in db.query(Service).all()}

        results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

        for row in rows[header_idx + 1:]:
            if not row: continue
            def get(field):
                idx = header_map.get(field)
                if idx is None or idx >= len(row): return None
                v = row[idx]
                return str(v).strip() if v is not None else None

            name = get("ten_dich_vu")
            if not name: continue

            nhom_name = get("nhom") or "Chưa phân loại"
            unit      = get("don_vi") or "Lần"
            try: price = float(get("don_gia_bv")) if get("don_gia_bv") else None
            except: price = None
            try: bhyt_price = float(get("don_gia_bh")) if get("don_gia_bh") else None
            except: bhyt_price = None
            try: order = int(float(get("thu_tu"))) if get("thu_tu") else 99
            except: order = 99

            # Tạo nhóm nếu chưa có
            nhom_key = nhom_name.strip().lower()
            if nhom_key not in groups_db:
                grp = ServiceGroup(name=nhom_name.strip(), display_order=len(groups_db)+1)
                db.add(grp); db.flush()
                groups_db[nhom_key] = grp

            grp_id = groups_db[nhom_key].id
            name_key = name.strip().lower()

            if name_key in svcs_db:
                # Cập nhật nếu đã tồn tại
                svc = svcs_db[name_key]
                if unit:       svc.unit       = unit
                if price:      svc.price      = price
                if bhyt_price is not None: svc.bhyt_price = bhyt_price
                results["updated"] += 1
            else:
                # Tạo mới
                svc = Service(
                    name=name.strip(), unit=unit, price=price,
                    bhyt_price=bhyt_price, group_id=grp_id,
                    display_order=order,
                )
                db.add(svc)
                svcs_db[name_key] = svc
                results["created"] += 1

        db.commit()
        return results

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Lỗi xử lý file: {str(e)}")
    finally:
        os.unlink(tmp_path)


@app.get("/api/services/export-excel")
def export_services_excel(db: Session = Depends(get_db)):
    """Xuất danh sách dịch vụ ra file Excel (dùng làm template)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dịch vụ"

    headers = ["Tên dịch vụ", "Nhóm", "Đơn vị", "Đơn giá BV", "Đơn giá BHYT", "Thứ tự"]
    col_widths = [50, 25, 10, 15, 15, 8]

    # Header row
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # Data
    groups = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    row = 2
    for grp in groups:
        for svc in sorted(grp.services, key=lambda x: x.display_order):
            ws.cell(row=row, column=1, value=svc.name)
            ws.cell(row=row, column=2, value=grp.name)
            ws.cell(row=row, column=3, value=svc.unit)
            ws.cell(row=row, column=4, value=svc.price)
            ws.cell(row=row, column=5, value=svc.bhyt_price)
            ws.cell(row=row, column=6, value=svc.display_order)
            # Format number cells
            for ci in [4, 5]:
                ws.cell(row=row, column=ci).number_format = '#,##0'
            row += 1

    # Alternate row colors
    from openpyxl.styles import PatternFill as PF
    for r in range(2, row):
        fill_color = "F7F4EF" if r % 2 == 0 else "FFFFFF"
        for ci in range(1, 7):
            ws.cell(row=r, column=ci).fill = PF("solid", fgColor=fill_color)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out = tmp.name
    wb.save(out)
    return FileResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename="danh_muc_dich_vu.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# GÓI KHÁM (EXAM PACKAGES)
# ══════════════════════════════════════════════════════════════════════════════
from models import ExamPackage, PackageService, PackagePatient

# ── Schemas ───────────────────────────────────────────────────────────────────
class PackageCreate(BaseModel):
    name: str
    description: Optional[str] = None

class PackageUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

class PackageServiceAdd(BaseModel):
    service_id: int
    so_luong: float = 1
    don_gia_bv: Optional[float] = None  # None = dùng giá từ danh mục
    don_gia_bh: Optional[float] = None

class PackageServiceUpdate(BaseModel):
    so_luong: Optional[float] = None
    don_gia_bv: Optional[float] = None
    don_gia_bh: Optional[float] = None

class PackagePatientAdd(BaseModel):
    patient_ids: List[int]
    ngay_kham: Optional[str] = None

class PackageBulkExport(BaseModel):
    format: str = "pdf"  # "pdf" or "word"
    patient_ids: Optional[List[int]] = None  # None = all patients in package

# ── Helper: build bill data from package ─────────────────────────────────────
def _package_bill_data(pkg: ExamPackage, patient: Patient,
                       ngay_kham: str, db) -> dict:
    """Build bill dict from a package + patient."""
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    grp_map   = {g.id: g for g in groups_db}

    items_by_group = defaultdict(list)
    tong_cong = 0.0

    for ps in pkg.package_services:
        svc = ps.service
        if not svc: continue
        price = ps.don_gia_bv if ps.don_gia_bv is not None else (svc.price or 0)
        dgbh  = ps.don_gia_bh if ps.don_gia_bh is not None else (svc.bhyt_price or 0)
        sl    = ps.so_luong or 1
        tt_bv = price * sl
        tong_cong += tt_bv
        items_by_group[svc.group_id].append({
            "name": svc.name, "unit": svc.unit,
            "so_luong": sl, "don_gia_bv": price,
            "don_gia_bh": dgbh, "ty_le_dv": 100,
            "thanh_tien_bv": tt_bv,
            "ty_le_bh": "", "thanh_tien_bh": None,
            "quy_bhyt": None, "nb_cung_tt": None,
            "khac": None, "nb_tu_tra": tt_bv,
        })

    groups_result = []
    for g in groups_db:
        if g.id in items_by_group:
            items = items_by_group[g.id]
            grp_total = sum(i["thanh_tien_bv"] for i in items)
            groups_result.append({
                "group_name": g.name,
                "group_prefix": GROUP_PREFIX_MAP.get(g.name, ""),
                "items": items,
                "tong_tt_bv": grp_total,
                "tong_bh": None, "tong_quy": None,
                "tong_nb_tu_tra": grp_total,
            })

    return {
        "ma_bn": patient.ma_bn, "so_kham_benh": "", "khoa": "Khoa Khám Bệnh",
        "ho_ten": patient.ho_ten, "gioi_tinh": patient.gioi_tinh,
        "ngay_sinh": patient.ngay_sinh, "dia_chi": patient.dia_chi,
        "ma_khu_vuc": "", "ma_the_bhyt": "", "bhyt_tu": "", "bhyt_den": "",
        "csdk_bhyt": "", "ma_csdk": "", "ngay_kham": ngay_kham or "",
        "chan_doan": "", "ma_benh": "", "benh_kem_theo": "", "ma_benh_kem": "",
        "muc_huong_bhyt": "", "ngay_tinh_tu": "", "ngay_tinh_den": "",
        "ten_goi_kham": pkg.name,
        "groups": groups_result,
        "tong_cong_bv": tong_cong, "tong_bh": None, "tong_quy": None,
        "tong_nb_tu_tra": tong_cong, "tong_chi_phi": tong_cong,
        "quy_bhyt_tt": 0, "nb_tu_tra": tong_cong, "khac_tt": 0,
        "total": tong_cong,
    }

def _pkg_to_dict(pkg: ExamPackage, db) -> dict:
    groups_db = db.query(ServiceGroup).order_by(ServiceGroup.display_order).all()
    grp_map = {g.id: g for g in groups_db}
    services = []
    for ps in sorted(pkg.package_services, key=lambda x: x.id):
        svc = ps.service
        if not svc: continue
        grp = grp_map.get(svc.group_id)
        services.append({
            "id": ps.id, "service_id": svc.id,
            "name": svc.name, "unit": svc.unit,
            "group_name": grp.name if grp else "",
            "so_luong": ps.so_luong,
            "don_gia_bv": ps.don_gia_bv if ps.don_gia_bv is not None else svc.price,
            "don_gia_bh": ps.don_gia_bh if ps.don_gia_bh is not None else svc.bhyt_price,
            "catalog_price": svc.price,
            "catalog_bhyt_price": svc.bhyt_price,
        })
    patients = []
    for pp in sorted(pkg.package_patients, key=lambda x: x.added_at):
        p = pp.patient
        if not p: continue
        patients.append({
            "id": pp.id, "patient_id": p.id,
            "ma_bn": p.ma_bn, "ho_ten": p.ho_ten,
            "gioi_tinh": p.gioi_tinh, "ngay_sinh": p.ngay_sinh,
            "dia_chi": p.dia_chi, "ngay_kham": pp.ngay_kham,
            "bill_id": pp.bill_id,
        })
    total_price = sum(
        (ps.don_gia_bv if ps.don_gia_bv is not None else (ps.service.price or 0)) * (ps.so_luong or 1)
        for ps in pkg.package_services if ps.service
    )
    return {
        "id": pkg.id, "name": pkg.name, "description": pkg.description,
        "created_at": pkg.created_at.isoformat(),
        "services": services, "patients": patients,
        "service_count": len(services), "patient_count": len(patients),
        "total_price": total_price,
    }

# ── CRUD Gói khám ─────────────────────────────────────────────────────────────
@app.get("/api/packages")
def list_packages(db: Session = Depends(get_db)):
    pkgs = db.query(ExamPackage).order_by(ExamPackage.created_at.desc()).all()
    return [_pkg_to_dict(p, db) for p in pkgs]

@app.post("/api/packages")
def create_package(body: PackageCreate, db: Session = Depends(get_db)):
    pkg = ExamPackage(name=body.name.strip(), description=body.description)
    db.add(pkg); db.commit(); db.refresh(pkg)
    return _pkg_to_dict(pkg, db)

@app.get("/api/packages/{pkg_id}")
def get_package(pkg_id: int, db: Session = Depends(get_db)):
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404, "Không tìm thấy gói khám")
    return _pkg_to_dict(pkg, db)

@app.put("/api/packages/{pkg_id}")
def update_package(pkg_id: int, body: PackageUpdate, db: Session = Depends(get_db)):
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404, "Không tìm thấy gói khám")
    if body.name is not None: pkg.name = body.name.strip()
    if body.description is not None: pkg.description = body.description
    db.commit(); db.refresh(pkg)
    return _pkg_to_dict(pkg, db)

@app.delete("/api/packages/{pkg_id}")
def delete_package(pkg_id: int, db: Session = Depends(get_db)):
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404, "Không tìm thấy gói khám")
    db.delete(pkg); db.commit()
    return {"ok": True}

# ── Dịch vụ trong gói ────────────────────────────────────────────────────────
@app.post("/api/packages/{pkg_id}/services")
def add_service_to_package(pkg_id: int, body: PackageServiceAdd,
                            db: Session = Depends(get_db)):
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404)
    svc = db.query(Service).filter(Service.id == body.service_id).first()
    if not svc: raise HTTPException(404, "Không tìm thấy dịch vụ")
    # Nếu đã có → cập nhật
    existing = next((ps for ps in pkg.package_services
                     if ps.service_id == body.service_id), None)
    if existing:
        existing.so_luong   = body.so_luong
        if body.don_gia_bv is not None: existing.don_gia_bv = body.don_gia_bv
        if body.don_gia_bh is not None: existing.don_gia_bh = body.don_gia_bh
    else:
        db.add(PackageService(
            package_id=pkg_id, service_id=body.service_id,
            so_luong=body.so_luong,
            don_gia_bv=body.don_gia_bv, don_gia_bh=body.don_gia_bh,
        ))
    db.commit(); db.refresh(pkg)
    return _pkg_to_dict(pkg, db)

@app.post("/api/packages/{pkg_id}/services/bulk")
def add_services_bulk(pkg_id: int, body: dict, db: Session = Depends(get_db)):
    """Thêm nhiều dịch vụ cùng lúc: {service_ids: [...]}"""
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404)
    existing_ids = {ps.service_id for ps in pkg.package_services}
    for sid in body.get("service_ids", []):
        svc = db.query(Service).filter(Service.id == sid).first()
        if not svc: continue
        if sid not in existing_ids:
            db.add(PackageService(package_id=pkg_id, service_id=sid, so_luong=1))
    db.commit(); db.refresh(pkg)
    return _pkg_to_dict(pkg, db)

@app.put("/api/packages/{pkg_id}/services/{ps_id}")
def update_package_service(pkg_id: int, ps_id: int, body: PackageServiceUpdate,
                            db: Session = Depends(get_db)):
    ps = db.query(PackageService).filter(
        PackageService.id == ps_id, PackageService.package_id == pkg_id).first()
    if not ps: raise HTTPException(404)
    if body.so_luong  is not None: ps.so_luong   = body.so_luong
    if body.don_gia_bv is not None: ps.don_gia_bv = body.don_gia_bv
    if body.don_gia_bh is not None: ps.don_gia_bh = body.don_gia_bh
    db.commit()
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    return _pkg_to_dict(pkg, db)

@app.delete("/api/packages/{pkg_id}/services/{ps_id}")
def remove_service_from_package(pkg_id: int, ps_id: int, db: Session = Depends(get_db)):
    ps = db.query(PackageService).filter(
        PackageService.id == ps_id, PackageService.package_id == pkg_id).first()
    if not ps: raise HTTPException(404)
    db.delete(ps); db.commit()
    return {"ok": True}

# ── Bệnh nhân trong gói ───────────────────────────────────────────────────────
@app.post("/api/packages/{pkg_id}/patients")
def add_patients_to_package(pkg_id: int, body: PackagePatientAdd,
                             db: Session = Depends(get_db)):
    """Thêm 1 hoặc nhiều bệnh nhân vào gói (tạo bảng kê ngay)."""
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404, "Không tìm thấy gói khám")
    existing_pids = {pp.patient_id for pp in pkg.package_patients}
    added = []
    for pid in body.patient_ids:
        patient = db.query(Patient).filter(Patient.id == pid).first()
        if not patient or pid in existing_pids: continue
        # Tạo bảng kê từ gói
        bill_data = _package_bill_data(pkg, patient, body.ngay_kham or "", db)
        bill = OutpatientBill(patient_id=patient.id, ngay_kham=body.ngay_kham or "")
        db.add(bill); db.flush()
        svc_map = {s.name.strip().lower(): s for s in db.query(Service).all()}
        for grp in bill_data["groups"]:
            for item in grp["items"]:
                svc_obj = svc_map.get(item["name"].strip().lower())
                tt_bv, _, _, _, nb_tu = calc_bhyt(
                    item["so_luong"], item["don_gia_bv"], item.get("don_gia_bh", 0),
                    100, None)
                db.add(BillItem(
                    bill_id=bill.id,
                    service_id=svc_obj.id if svc_obj else None,
                    so_luong=item["so_luong"],
                    don_gia_bv=item["don_gia_bv"],
                    don_gia_bh=item.get("don_gia_bh", 0),
                    ty_le_tt_dv=100, thanh_tien_bv=tt_bv, nb_tu_tra=nb_tu,
                ))
        pp = PackagePatient(package_id=pkg_id, patient_id=pid,
                            bill_id=bill.id, ngay_kham=body.ngay_kham)
        db.add(pp)
        existing_pids.add(pid)
        added.append({"patient_id": pid, "bill_id": bill.id})
    db.commit(); db.refresh(pkg)
    return {"added": added, "package": _pkg_to_dict(pkg, db)}

@app.delete("/api/packages/{pkg_id}/patients/{pp_id}")
def remove_patient_from_package(pkg_id: int, pp_id: int,
                                 delete_bill: bool = False,
                                 db: Session = Depends(get_db)):
    pp = db.query(PackagePatient).filter(
        PackagePatient.id == pp_id, PackagePatient.package_id == pkg_id).first()
    if not pp: raise HTTPException(404)
    if delete_bill and pp.bill_id:
        bill = db.query(OutpatientBill).filter(OutpatientBill.id == pp.bill_id).first()
        if bill: db.delete(bill)
    db.delete(pp); db.commit()
    return {"ok": True}

# ── Import gói khám từ Excel ──────────────────────────────────────────────────
@app.post("/api/packages/{pkg_id}/import-services-excel")
async def import_package_services_excel(pkg_id: int, file: UploadFile = File(...),
                                         db: Session = Depends(get_db)):
    """Import danh sách dịch vụ vào gói từ file Excel."""
    import openpyxl
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read()); tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active; rows = list(ws.iter_rows(values_only=True))
        # Find header
        header_idx = None; name_col = qty_col = price_col = bhyt_col = None
        for i, row in enumerate(rows[:5]):
            if not row: continue
            rl = [str(c).strip().lower() if c else "" for c in row]
            for j, h in enumerate(rl):
                if h in ["tên dịch vụ","ten dich vu","dịch vụ","name","service"]: name_col=j
                if h in ["số lượng","so luong","sl","qty","quantity"]: qty_col=j
                if h in ["đơn giá bv","don gia bv","giá","gia","price"]: price_col=j
                if h in ["đơn giá bhyt","don gia bhyt","bhyt price"]: bhyt_col=j
            if name_col is not None: header_idx=i; break

        if header_idx is None: raise HTTPException(400, "Không tìm thấy cột 'Tên dịch vụ'")

        svc_map = {s.name.strip().lower(): s for s in db.query(Service).all()}
        existing_ids = {ps.service_id for ps in pkg.package_services}
        added = updated = skipped = 0

        for row in rows[header_idx+1:]:
            if not row or not row[name_col]: continue
            name = str(row[name_col]).strip()
            svc  = svc_map.get(name.lower())
            if not svc: skipped += 1; continue

            qty   = int(float(row[qty_col]))   if qty_col  and row[qty_col]  else 1
            price = float(row[price_col])       if price_col and row[price_col] else None
            bhyt  = float(row[bhyt_col])        if bhyt_col  and row[bhyt_col]  else None

            existing = next((ps for ps in pkg.package_services
                             if ps.service_id == svc.id), None)
            if existing:
                existing.so_luong = qty
                if price is not None: existing.don_gia_bv = price
                if bhyt  is not None: existing.don_gia_bh = bhyt
                updated += 1
            elif svc.id not in existing_ids:
                db.add(PackageService(
                    package_id=pkg_id, service_id=svc.id, so_luong=qty,
                    don_gia_bv=price, don_gia_bh=bhyt,
                ))
                existing_ids.add(svc.id); added += 1

        db.commit(); db.refresh(pkg)
        return {"added": added, "updated": updated, "skipped": skipped,
                "package": _pkg_to_dict(pkg, db)}
    finally:
        os.unlink(tmp_path)

# ── Import bệnh nhân hàng loạt vào gói ───────────────────────────────────────
@app.post("/api/packages/{pkg_id}/import-patients-excel")
async def import_package_patients_excel(pkg_id: int, file: UploadFile = File(...),
                                         ngay_kham: str = "",
                                         db: Session = Depends(get_db)):
    """Import danh sách bệnh nhân (từ mã BN hoặc họ tên) vào gói."""
    import openpyxl
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read()); tmp_path = tmp.name
    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active; rows = list(ws.iter_rows(values_only=True))
        header_idx = None; mabn_col = None
        for i, row in enumerate(rows[:5]):
            if not row: continue
            rl = [str(c).strip().lower() if c else "" for c in row]
            for j, h in enumerate(rl):
                if h in ["mã bn","ma bn","id","mã số","patient id"]:
                    mabn_col = j; header_idx = i; break
            if header_idx is not None: break
        if header_idx is None or mabn_col is None:
            raise HTTPException(400, "Không tìm thấy cột 'Mã BN'")

        existing_pids = {pp.patient_id for pp in pkg.package_patients}
        added = skipped = not_found = 0

        all_pids = []
        for row in rows[header_idx+1:]:
            if not row or not row[mabn_col]: continue
            ma_bn = str(row[mabn_col]).strip()
            patient = db.query(Patient).filter(Patient.ma_bn == ma_bn).first()
            if not patient: not_found += 1; continue
            if patient.id in existing_pids: skipped += 1; continue
            all_pids.append(patient.id)
            existing_pids.add(patient.id)

        # Tạo bảng kê cho từng bệnh nhân
        for pid in all_pids:
            patient = db.query(Patient).filter(Patient.id == pid).first()
            bill_data = _package_bill_data(pkg, patient, ngay_kham, db)
            bill = OutpatientBill(patient_id=patient.id, ngay_kham=ngay_kham)
            db.add(bill); db.flush()
            svc_map_l = {s.name.strip().lower(): s for s in db.query(Service).all()}
            for grp in bill_data["groups"]:
                for item in grp["items"]:
                    so = svc_map_l.get(item["name"].strip().lower())
                    tt_bv, _, _, _, nb_tu = calc_bhyt(
                        item["so_luong"], item["don_gia_bv"], item.get("don_gia_bh",0), 100, None)
                    db.add(BillItem(bill_id=bill.id,
                        service_id=so.id if so else None,
                        so_luong=item["so_luong"], don_gia_bv=item["don_gia_bv"],
                        don_gia_bh=item.get("don_gia_bh",0),
                        ty_le_tt_dv=100, thanh_tien_bv=tt_bv, nb_tu_tra=nb_tu))
            db.add(PackagePatient(package_id=pkg_id, patient_id=pid,
                                  bill_id=bill.id, ngay_kham=ngay_kham))
            added += 1

        db.commit(); db.refresh(pkg)
        return {"added": added, "skipped": skipped, "not_found": not_found,
                "package": _pkg_to_dict(pkg, db)}
    finally:
        os.unlink(tmp_path)

# ── Xuất bảng kê theo gói ─────────────────────────────────────────────────────
@app.post("/api/packages/{pkg_id}/export")
def export_package(pkg_id: int, body: PackageBulkExport,
                   db: Session = Depends(get_db)):
    """Xuất bảng kê của tất cả hoặc một số bệnh nhân trong gói → ZIP."""
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404)

    pp_list = pkg.package_patients
    if body.patient_ids:
        pp_list = [pp for pp in pp_list if pp.patient_id in set(body.patient_ids)]

    if not pp_list: raise HTTPException(400, "Không có bệnh nhân nào để xuất")

    with tempfile.TemporaryDirectory() as tmp:
        zip_path = os.path.join(tmp, "export.zip")
        errors = []
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for pp in pp_list:
                patient = pp.patient
                if not patient: continue
                try:
                    bill_data = _package_bill_data(
                        pkg, patient, pp.ngay_kham or "", db)
                    ext  = "docx" if body.format == "word" else "pdf"
                    fname = f"bang_ke_{patient.ma_bn}_{patient.ho_ten.replace(' ','_')}.{ext}"
                    out  = os.path.join(tmp, fname)
                    if body.format == "word":
                        generate_word(bill_data, out)
                    else:
                        generate_pdf(bill_data, out)
                    zf.write(out, fname)
                except Exception as e:
                    errors.append(f"{patient.ma_bn}: {str(e)}")

        with open(zip_path, 'rb') as f:
            zip_b64 = base64.b64encode(f.read()).decode()

        return {
            "zip_base64": zip_b64,
            "filename": f"goi_kham_{pkg.name.replace(' ','_')}.zip",
            "success_count": len(pp_list) - len(errors),
            "errors": errors,
        }

# ── Áp dụng dịch vụ gói cho bệnh nhân đã có trong hệ thống ──────────────────
@app.post("/api/packages/{pkg_id}/apply-services")
def apply_package_services(pkg_id: int, body: dict, db: Session = Depends(get_db)):
    """
    Áp dụng dịch vụ mới/thêm từ gói vào bảng kê của bệnh nhân đã có.
    body: {
      patient_ids: [int],           # danh sách patient_id
      service_ids: [int] | null,    # null = áp dụng tất cả DV trong gói
      replace: bool                 # true = xóa DV cũ rồi thêm mới
    }
    """
    pkg = db.query(ExamPackage).filter(ExamPackage.id == pkg_id).first()
    if not pkg: raise HTTPException(404)

    pid_list   = body.get("patient_ids", [])
    svc_filter = set(body.get("service_ids") or [])
    replace    = body.get("replace", False)

    results = []
    for pid in pid_list:
        # Tìm bảng kê mới nhất của bệnh nhân này trong gói
        pp = db.query(PackagePatient).filter(
            PackagePatient.package_id == pkg_id,
            PackagePatient.patient_id == pid,
        ).order_by(PackagePatient.added_at.desc()).first()

        if not pp or not pp.bill_id:
            results.append({"patient_id": pid, "status": "no_bill"})
            continue

        bill = db.query(OutpatientBill).filter(OutpatientBill.id == pp.bill_id).first()
        if not bill:
            results.append({"patient_id": pid, "status": "bill_not_found"})
            continue

        if replace:
            # Xóa dịch vụ cũ khớp với gói
            pkg_svc_ids = {ps.service_id for ps in pkg.package_services}
            to_delete = svc_filter if svc_filter else pkg_svc_ids
            for item in list(bill.items):
                if item.service_id in to_delete:
                    db.delete(item)

        # Thêm dịch vụ từ gói
        svc_map = {s.name.strip().lower(): s for s in db.query(Service).all()}
        for ps in pkg.package_services:
            if svc_filter and ps.service_id not in svc_filter:
                continue
            svc = ps.service
            if not svc: continue
            price = ps.don_gia_bv if ps.don_gia_bv is not None else (svc.price or 0)
            dgbh  = ps.don_gia_bh if ps.don_gia_bh is not None else (svc.bhyt_price or 0)
            tt_bv = price * (ps.so_luong or 1)
            db.add(BillItem(
                bill_id=bill.id, service_id=svc.id,
                so_luong=ps.so_luong or 1,
                don_gia_bv=price, don_gia_bh=dgbh,
                ty_le_tt_dv=100, thanh_tien_bv=tt_bv, nb_tu_tra=tt_bv,
            ))
        results.append({"patient_id": pid, "bill_id": bill.id, "status": "ok"})

    db.commit()
    return {"results": results}


# ══════════════════════════════════════════════════════════════════════════════
# BỆNH NHÂN - CRUD ĐẦY ĐỦ
# ══════════════════════════════════════════════════════════════════════════════
class PatientCreate(BaseModel):
    ma_bn: str
    ho_ten: str
    gioi_tinh: Optional[str] = "Nam"
    ngay_sinh: Optional[str] = ""
    dia_chi: Optional[str] = ""
    ten_goi_kham: Optional[str] = ""

class PatientUpdate(BaseModel):
    ho_ten: Optional[str] = None
    gioi_tinh: Optional[str] = None
    ngay_sinh: Optional[str] = None
    dia_chi: Optional[str] = None

@app.get("/api/patients")
def list_patients(db: Session = Depends(get_db)):
    patients = db.query(Patient).order_by(Patient.created_at.desc()).all()
    return [{
        "id": p.id, "ma_bn": p.ma_bn, "ho_ten": p.ho_ten,
        "gioi_tinh": p.gioi_tinh, "ngay_sinh": p.ngay_sinh,
        "dia_chi": p.dia_chi, "so_lan_kham": len(p.bills),
        "created_at": p.created_at.isoformat(),
    } for p in patients]

@app.post("/api/patients")
def create_patient(body: PatientCreate, db: Session = Depends(get_db)):
    existing = db.query(Patient).filter(Patient.ma_bn == body.ma_bn).first()
    if existing:
        raise HTTPException(400, f"Mã BN '{body.ma_bn}' đã tồn tại")
    p = Patient(
        ma_bn=body.ma_bn.strip(), ho_ten=body.ho_ten.strip(),
        gioi_tinh=body.gioi_tinh, ngay_sinh=body.ngay_sinh,
        dia_chi=body.dia_chi, ten_goi_kham=body.ten_goi_kham,
    )
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id, "ma_bn": p.ma_bn, "ho_ten": p.ho_ten,
            "gioi_tinh": p.gioi_tinh, "ngay_sinh": p.ngay_sinh,
            "dia_chi": p.dia_chi, "so_lan_kham": 0}

@app.put("/api/patients/{patient_id}")
def update_patient(patient_id: int, body: PatientUpdate, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == patient_id).first()
    if not p: raise HTTPException(404, "Không tìm thấy bệnh nhân")
    if body.ho_ten    is not None: p.ho_ten    = body.ho_ten
    if body.gioi_tinh is not None: p.gioi_tinh = body.gioi_tinh
    if body.ngay_sinh is not None: p.ngay_sinh = body.ngay_sinh
    if body.dia_chi   is not None: p.dia_chi   = body.dia_chi
    db.commit(); db.refresh(p)
    return {"id": p.id, "ma_bn": p.ma_bn, "ho_ten": p.ho_ten,
            "gioi_tinh": p.gioi_tinh, "ngay_sinh": p.ngay_sinh,
            "dia_chi": p.dia_chi, "so_lan_kham": len(p.bills)}

@app.delete("/api/patients/{patient_id}")
def delete_patient(patient_id: int, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == patient_id).first()
    if not p: raise HTTPException(404, "Không tìm thấy bệnh nhân")
    db.delete(p); db.commit()
    return {"ok": True}


# ══════════════════════════════════════════════════════════════════════════════
# BẢNG KÊ - TẠO TAY
# ══════════════════════════════════════════════════════════════════════════════
class BillCreate(BaseModel):
    patient_id: int
    ngay_kham: Optional[str] = ""

@app.post("/api/bills")
def create_bill(body: BillCreate, db: Session = Depends(get_db)):
    """Tạo bảng kê trống cho bệnh nhân đã có."""
    p = db.query(Patient).filter(Patient.id == body.patient_id).first()
    if not p: raise HTTPException(404, "Không tìm thấy bệnh nhân")
    bill = OutpatientBill(patient_id=p.id, ngay_kham=body.ngay_kham or "")
    db.add(bill); db.commit(); db.refresh(bill)
    return {
        "id": bill.id, "ma_bn": p.ma_bn, "ho_ten": p.ho_ten,
        "ngay_kham": bill.ngay_kham, "total": 0, "items_count": 0,
        "created_at": bill.created_at.isoformat(),
    }


# ══════════════════════════════════════════════════════════════════════════════
# QUẢN LÝ MẪU IN (TEMPLATES)
# ══════════════════════════════════════════════════════════════════════════════
import json as _json

TEMPLATES_DIR = "/app/templates"
TEMPLATES_META = "/app/templates/templates_meta.json"

def _load_meta() -> dict:
    """Load metadata của tất cả templates."""
    if os.path.exists(TEMPLATES_META):
        try:
            with open(TEMPLATES_META) as f:
                return _json.load(f)
        except: pass
    # Tự động scan templates có sẵn
    meta = {}
    if os.path.isdir(TEMPLATES_DIR):
        for fname in os.listdir(TEMPLATES_DIR):
            if fname.endswith('.docx'):
                key = fname
                meta[key] = {
                    "filename": fname,
                    "name": fname.replace('.docx','').replace('_',' ').replace('-',' '),
                    "description": "",
                    "is_default": fname == "template.docx",
                    "created_at": "",
                }
    return meta

def _save_meta(meta: dict):
    with open(TEMPLATES_META, 'w', encoding='utf-8') as f:
        _json.dump(meta, f, ensure_ascii=False, indent=2)

def _get_default_template() -> str:
    """Trả về đường dẫn template mặc định."""
    meta = _load_meta()
    for key, info in meta.items():
        if info.get("is_default"):
            path = os.path.join(TEMPLATES_DIR, info["filename"])
            if os.path.exists(path):
                return path
    # Fallback: tìm file .docx bất kỳ
    if os.path.isdir(TEMPLATES_DIR):
        for f in os.listdir(TEMPLATES_DIR):
            if f.endswith('.docx'):
                return os.path.join(TEMPLATES_DIR, f)
    raise FileNotFoundError("Không tìm thấy template nào trong /app/templates/")

@app.get("/api/templates")
def list_templates():
    """Danh sách tất cả templates."""
    meta = _load_meta()
    result = []
    for key, info in meta.items():
        path = os.path.join(TEMPLATES_DIR, info["filename"])
        size = os.path.getsize(path) if os.path.exists(path) else 0
        result.append({
            "id": key,
            "filename": info["filename"],
            "name": info["name"],
            "description": info.get("description",""),
            "is_default": info.get("is_default", False),
            "size_kb": round(size/1024, 1),
            "created_at": info.get("created_at",""),
        })
    return sorted(result, key=lambda x: (not x["is_default"], x["name"]))

@app.post("/api/templates/upload")
async def upload_template(
    file: UploadFile = File(...),
    name: str = "",
    description: str = "",
    set_default: bool = False,
):
    """Upload file Word (.docx) làm template mới."""
    if not file.filename.endswith('.docx'):
        raise HTTPException(400, "Chỉ chấp nhận file .docx")

    # Lưu file
    safe_name = file.filename.replace(' ', '_')
    dest = os.path.join(TEMPLATES_DIR, safe_name)
    content = await file.read()
    with open(dest, 'wb') as f:
        f.write(content)

    # Cập nhật metadata
    meta = _load_meta()
    display_name = name.strip() or safe_name.replace('.docx','').replace('_',' ')
    if set_default:
        for k in meta: meta[k]["is_default"] = False
    meta[safe_name] = {
        "filename": safe_name,
        "name": display_name,
        "description": description,
        "is_default": set_default or len(meta) == 0,
        "created_at": __import__('datetime').datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    _save_meta(meta)
    return {"ok": True, "filename": safe_name, "name": display_name}

@app.put("/api/templates/{template_id}")
def update_template_meta(template_id: str, body: dict):
    """Cập nhật tên, mô tả, đặt mặc định."""
    meta = _load_meta()
    if template_id not in meta:
        raise HTTPException(404, "Không tìm thấy template")
    if "name" in body:        meta[template_id]["name"]        = body["name"]
    if "description" in body: meta[template_id]["description"] = body["description"]
    if body.get("is_default"):
        for k in meta: meta[k]["is_default"] = False
        meta[template_id]["is_default"] = True
    _save_meta(meta)
    return {"ok": True}

@app.delete("/api/templates/{template_id}")
def delete_template(template_id: str):
    """Xóa template (không được xóa template mặc định duy nhất)."""
    meta = _load_meta()
    if template_id not in meta:
        raise HTTPException(404, "Không tìm thấy template")
    info = meta[template_id]
    # Không cho xóa nếu là mặc định và chỉ còn 1
    if info.get("is_default") and len(meta) == 1:
        raise HTTPException(400, "Không thể xóa template mặc định duy nhất")
    # Xóa file
    path = os.path.join(TEMPLATES_DIR, info["filename"])
    if os.path.exists(path): os.remove(path)
    del meta[template_id]
    # Nếu xóa default → đặt cái đầu tiên làm default
    if info.get("is_default") and meta:
        first = next(iter(meta))
        meta[first]["is_default"] = True
    _save_meta(meta)
    return {"ok": True}

@app.get("/api/templates/{template_id}/download")
def download_template(template_id: str):
    """Tải về file template để xem/chỉnh sửa."""
    meta = _load_meta()
    if template_id not in meta:
        raise HTTPException(404)
    path = os.path.join(TEMPLATES_DIR, meta[template_id]["filename"])
    if not os.path.exists(path):
        raise HTTPException(404, "File không tồn tại")
    return FileResponse(path,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=meta[template_id]["filename"])

# ── Override export endpoints để hỗ trợ chọn template ──────────────────────
@app.get("/api/bills/{bill_id}/export/word")
def export_word_v2(bill_id: int, template_id: str = None,
                   db: Session = Depends(get_db)):
    bill_data = get_bill(bill_id, db)
    # Chọn template
    if template_id:
        meta = _load_meta()
        if template_id in meta:
            tpl_path = os.path.join(TEMPLATES_DIR, meta[template_id]["filename"])
        else:
            tpl_path = _get_default_template()
    else:
        tpl_path = _get_default_template()

    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        out = tmp.name
    try:
        import word_generator as _wg
        _wg._OVERRIDE_TEMPLATE = tpl_path
        generate_word(bill_data, out)
        _wg._OVERRIDE_TEMPLATE = None
        fname = f"bang_ke_{bill_data['ma_bn']}_{bill_data['ho_ten'].replace(' ','_')}.docx"
        return FileResponse(out,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename=fname)
    except Exception as e:
        os.unlink(out); raise HTTPException(500, str(e))

@app.get("/api/bills/{bill_id}/export/pdf")
def export_pdf_v2(bill_id: int, template_id: str = None,
                  db: Session = Depends(get_db)):
    bill_data = get_bill(bill_id, db)
    if template_id:
        meta = _load_meta()
        tpl_path = os.path.join(TEMPLATES_DIR, meta[template_id]["filename"]) \
                   if template_id in meta else _get_default_template()
    else:
        tpl_path = _get_default_template()

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        out = tmp.name
    try:
        import word_generator as _wg
        _wg._OVERRIDE_TEMPLATE = tpl_path
        generate_pdf(bill_data, out)
        _wg._OVERRIDE_TEMPLATE = None
        fname = f"bang_ke_{bill_data['ma_bn']}_{bill_data['ho_ten'].replace(' ','_')}.pdf"
        return FileResponse(out, media_type="application/pdf", filename=fname)
    except Exception as e:
        os.unlink(out); raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════════════════════
# AUTH & USER MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════
from jose import JWTError, jwt as _jwt
from passlib.context import CryptContext
from fastapi import Request, Cookie
from fastapi.responses import JSONResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from models import User, ActivityLog
from datetime import timedelta

SECRET_KEY = os.getenv("JWT_SECRET", "hongduc2-super-secret-key-2026")
ALGORITHM  = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 8  # 8 giờ

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)

# ── Helpers ───────────────────────────────────────────────────────────────────
def verify_password(plain, hashed): return pwd_ctx.verify(plain[:72], hashed)
def hash_password(pw): return pwd_ctx.hash(pw[:72])  # bcrypt max 72 bytes

def create_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode["exp"] = expire
    return _jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_token_from_request(request: Request) -> Optional[str]:
    """Lấy token từ header Authorization hoặc cookie."""
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:]
    return request.cookies.get("access_token")

def get_current_user_optional(request: Request, db: Session = Depends(get_db)) -> Optional[User]:
    token = get_token_from_request(request)
    if not token: return None
    try:
        payload = _jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if not username: return None
        return db.query(User).filter(User.username == username, User.is_active == True).first()
    except JWTError:
        return None

def require_user(request: Request, db: Session = Depends(get_db)) -> User:
    user = get_current_user_optional(request, db)
    if not user: raise HTTPException(401, "Chưa đăng nhập")
    return user

def require_admin(request: Request, db: Session = Depends(get_db)) -> User:
    user = require_user(request, db)
    if user.role != "admin": raise HTTPException(403, "Chỉ admin mới có quyền này")
    return user

def log_activity(db: Session, user: Optional[User], action: str,
                 resource: str = None, resource_id: str = None,
                 detail: str = None, request: Request = None,
                 status: str = "success"):
    ip = ua = None
    if request:
        ip = request.client.host if request.client else None
        ua = request.headers.get("user-agent", "")[:300]
    entry = ActivityLog(
        user_id=user.id if user else None,
        username=user.username if user else "anonymous",
        action=action, resource=resource, resource_id=str(resource_id) if resource_id else None,
        detail=detail, ip_address=ip, user_agent=ua, status=status,
    )
    db.add(entry); db.commit()

# ── Auth Schemas ──────────────────────────────────────────────────────────────
class UserCreate(BaseModel):
    username: str
    password: str
    email: Optional[str] = None
    full_name: Optional[str] = None
    role: str = "user"

class UserUpdate(BaseModel):
    email: Optional[str] = None
    full_name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None

class LoginRequest(BaseModel):
    username: str
    password: str

# ── Auth Endpoints ────────────────────────────────────────────────────────────
@app.post("/api/auth/login")
async def login(body: LoginRequest, request: Request, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == body.username).first()
    if not user or not verify_password(body.password, user.hashed_pw):
        log_activity(db, None, "LOGIN_FAILED", detail=f"username={body.username}", request=request, status="error")
        raise HTTPException(401, "Sai tên đăng nhập hoặc mật khẩu")
    if not user.is_active:
        raise HTTPException(403, "Tài khoản đã bị khóa")

    token = create_token({"sub": user.username, "role": user.role})
    user.last_login = datetime.utcnow()
    db.commit()

    log_activity(db, user, "LOGIN", detail="Đăng nhập thành công", request=request)
    return {
        "access_token": token, "token_type": "bearer",
        "user": {"id": user.id, "username": user.username,
                 "full_name": user.full_name, "role": user.role, "email": user.email}
    }

@app.post("/api/auth/logout")
async def logout(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_optional(request, db)
    if user:
        log_activity(db, user, "LOGOUT", request=request)
    return {"ok": True}

@app.get("/api/auth/me")
async def get_me(request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    return {"id": user.id, "username": user.username,
            "full_name": user.full_name, "role": user.role,
            "email": user.email, "created_at": user.created_at.isoformat(),
            "last_login": user.last_login.isoformat() if user.last_login else None}

@app.put("/api/auth/change-password")
async def change_password(body: dict, request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    old_pw = body.get("old_password", "")
    new_pw = body.get("new_password", "")
    if not verify_password(old_pw, user.hashed_pw):
        raise HTTPException(400, "Mật khẩu cũ không đúng")
    if len(new_pw) < 6:
        raise HTTPException(400, "Mật khẩu mới phải ít nhất 6 ký tự")
    user.hashed_pw = hash_password(new_pw)
    db.commit()
    log_activity(db, user, "CHANGE_PASSWORD", request=request)
    return {"ok": True}

# ── User Management (Admin) ───────────────────────────────────────────────────
@app.get("/api/users")
async def list_users(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    users = db.query(User).order_by(User.created_at.desc()).all()
    return [{
        "id": u.id, "username": u.username, "email": u.email,
        "full_name": u.full_name, "role": u.role, "is_active": u.is_active,
        "created_at": u.created_at.isoformat(),
        "last_login": u.last_login.isoformat() if u.last_login else None,
    } for u in users]

@app.post("/api/users")
async def create_user(body: UserCreate, request: Request, db: Session = Depends(get_db)):
    admin = require_admin(request, db)
    if db.query(User).filter(User.username == body.username).first():
        raise HTTPException(400, f"Username '{body.username}' đã tồn tại")
    user = User(
        username=body.username, email=body.email, full_name=body.full_name,
        hashed_pw=hash_password(body.password), role=body.role,
    )
    db.add(user); db.commit(); db.refresh(user)
    log_activity(db, admin, "CREATE_USER", resource="users", resource_id=user.id,
                 detail=f"Tạo user {user.username}", request=request)
    return {"id": user.id, "username": user.username, "role": user.role}

@app.put("/api/users/{user_id}")
async def update_user(user_id: int, body: UserUpdate, request: Request, db: Session = Depends(get_db)):
    admin = require_admin(request, db)
    user = db.query(User).filter(User.id == user_id).first()
    if not user: raise HTTPException(404, "Không tìm thấy user")
    if body.email      is not None: user.email     = body.email
    if body.full_name  is not None: user.full_name  = body.full_name
    if body.role       is not None: user.role       = body.role
    if body.is_active  is not None: user.is_active  = body.is_active
    if body.password:               user.hashed_pw  = hash_password(body.password)
    db.commit()
    log_activity(db, admin, "UPDATE_USER", resource="users", resource_id=user_id,
                 detail=f"Cập nhật user {user.username}", request=request)
    return {"ok": True}

@app.delete("/api/users/{user_id}")
async def delete_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    admin = require_admin(request, db)
    user = db.query(User).filter(User.id == user_id).first()
    if not user: raise HTTPException(404)
    if user.username == "admin": raise HTTPException(400, "Không thể xóa admin chính")
    username = user.username
    db.delete(user); db.commit()
    log_activity(db, admin, "DELETE_USER", resource="users", resource_id=user_id,
                 detail=f"Xóa user {username}", request=request)
    return {"ok": True}

# ── Activity Logs (Admin) ─────────────────────────────────────────────────────
@app.get("/api/logs")
async def get_logs(
    request: Request, db: Session = Depends(get_db),
    page: int = 1, limit: int = 50,
    username: str = None, action: str = None, resource: str = None,
    date_from: str = None, date_to: str = None,
):
    require_admin(request, db)
    q = db.query(ActivityLog).order_by(ActivityLog.created_at.desc())
    if username: q = q.filter(ActivityLog.username.ilike(f"%{username}%"))
    if action:   q = q.filter(ActivityLog.action == action)
    if resource: q = q.filter(ActivityLog.resource == resource)
    if date_from:
        try:
            from datetime import datetime as dt
            q = q.filter(ActivityLog.created_at >= dt.strptime(date_from, "%Y-%m-%d"))
        except: pass
    if date_to:
        try:
            from datetime import datetime as dt
            q = q.filter(ActivityLog.created_at <= dt.strptime(date_to + " 23:59:59", "%Y-%m-%d %H:%M:%S"))
        except: pass

    total = q.count()
    logs  = q.offset((page-1)*limit).limit(limit).all()
    return {
        "total": total, "page": page, "limit": limit,
        "pages": (total + limit - 1) // limit,
        "logs": [{
            "id": l.id, "username": l.username, "action": l.action,
            "resource": l.resource, "resource_id": l.resource_id,
            "detail": l.detail, "ip_address": l.ip_address,
            "status": l.status,
            "created_at": l.created_at.isoformat(),
        } for l in logs]
    }

@app.get("/api/logs/stats")
async def get_log_stats(request: Request, db: Session = Depends(get_db)):
    require_admin(request, db)
    from sqlalchemy import func
    # Top actions
    top_actions = db.query(
        ActivityLog.action,
        func.count(ActivityLog.id).label("count")
    ).group_by(ActivityLog.action).order_by(func.count(ActivityLog.id).desc()).limit(10).all()

    # Top users
    top_users = db.query(
        ActivityLog.username,
        func.count(ActivityLog.id).label("count")
    ).group_by(ActivityLog.username).order_by(func.count(ActivityLog.id).desc()).limit(10).all()

    # Recent logins
    recent_logins = db.query(ActivityLog).filter(
        ActivityLog.action == "LOGIN"
    ).order_by(ActivityLog.created_at.desc()).limit(10).all()

    total_logs   = db.query(ActivityLog).count()
    total_users  = db.query(User).count()
    active_users = db.query(User).filter(User.is_active == True).count()

    return {
        "total_logs": total_logs, "total_users": total_users,
        "active_users": active_users,
        "top_actions": [{"action": a.action, "count": a.count} for a in top_actions],
        "top_users":   [{"username": u.username, "count": u.count} for u in top_users],
        "recent_logins": [{
            "username": l.username, "ip_address": l.ip_address,
            "created_at": l.created_at.isoformat()
        } for l in recent_logins],
    }

@app.delete("/api/logs/clear")
async def clear_old_logs(days: int = 30, request: Request = None, db: Session = Depends(get_db)):
    admin = require_admin(request, db)
    cutoff = datetime.utcnow() - timedelta(days=days)
    deleted = db.query(ActivityLog).filter(ActivityLog.created_at < cutoff).delete()
    db.commit()
    log_activity(db, admin, "CLEAR_LOGS", detail=f"Xóa log cũ hơn {days} ngày ({deleted} bản ghi)", request=request)
    return {"deleted": deleted}

# ── Middleware ghi log tự động ────────────────────────────────────────────────
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest

# Map method+path → action name
LOG_MAP = {
    ("POST",   "/api/bills"):            ("CREATE_BILL",   "bills"),
    ("DELETE", "/api/bills/"):           ("DELETE_BILL",   "bills"),
    ("POST",   "/api/bills/export-bulk"):("EXPORT_BULK",   "bills"),
    ("POST",   "/api/patients"):         ("CREATE_PATIENT","patients"),
    ("PUT",    "/api/patients/"):        ("UPDATE_PATIENT","patients"),
    ("DELETE", "/api/patients/"):        ("DELETE_PATIENT","patients"),
    ("POST",   "/api/services"):         ("CREATE_SERVICE","services"),
    ("PUT",    "/api/services/"):        ("UPDATE_SERVICE","services"),
    ("DELETE", "/api/services/"):        ("DELETE_SERVICE","services"),
    ("POST",   "/api/service-groups"):   ("CREATE_GROUP",  "service_groups"),
    ("DELETE", "/api/service-groups/"):  ("DELETE_GROUP",  "service_groups"),
    ("POST",   "/api/packages"):         ("CREATE_PACKAGE","packages"),
    ("DELETE", "/api/packages/"):        ("DELETE_PACKAGE","packages"),
    ("POST",   "/api/upload-excel"):     ("IMPORT_EXCEL",  "bills"),
    ("GET",    "/api/bills/"):           ("EXPORT_BILL",   "bills"),
    ("POST",   "/api/templates/upload"): ("UPLOAD_TEMPLATE","templates"),
    ("DELETE", "/api/templates/"):       ("DELETE_TEMPLATE","templates"),
}

class ActivityLogMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: StarletteRequest, call_next):
        response = await call_next(request)

        # Chỉ log các API write operations
        method = request.method
        path   = request.url.path

        if not path.startswith("/api/") or path.startswith("/api/auth/") or path.startswith("/api/logs"):
            return response
        if method not in ("POST", "PUT", "DELETE", "PATCH"):
            return response
        if response.status_code >= 400:
            return response

        # Tìm action phù hợp
        action = resource = None
        for (m, p), (a, r) in LOG_MAP.items():
            if method == m and path.startswith(p):
                action, resource = a, r
                break

        if action:
            try:
                from database import SessionLocal as _SL
                db2 = _SL()
                token = get_token_from_request(request)
                user = None
                if token:
                    try:
                        payload = _jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
                        uname = payload.get("sub")
                        if uname:
                            user = db2.query(User).filter(User.username == uname).first()
                    except: pass
                # Extract resource_id from path
                parts = path.rstrip('/').split('/')
                rid = parts[-1] if parts[-1].isdigit() else None
                log_activity(db2, user, action, resource=resource, resource_id=rid,
                             detail=f"{method} {path}", request=request)
                db2.close()
            except: pass

        return response

app.add_middleware(ActivityLogMiddleware)
